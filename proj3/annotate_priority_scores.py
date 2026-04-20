from __future__ import annotations

import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Any
from urllib import error, request

import openpyxl

ROOT = Path(__file__).resolve().parent
ENV_PATH = ROOT / ".env"
XLSX_PATH = ROOT / "第7组反馈.xlsx"
SHEET_NAME = "feedback"
HEADER_ROW = 2
DATA_START_ROW = 3

COL_EMOTION = "情绪强度"
COL_RISK = "风险等级"
COL_URGENCY = "紧急程度"
TARGET_COLS = [COL_EMOTION, COL_RISK, COL_URGENCY]


@dataclass
class LlmConfig:
    api_key: str
    base_url: str
    model: str


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def build_llm_config() -> LlmConfig:
    load_env_file(ENV_PATH)
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com").strip()
    model = os.getenv("DEEPSEEK_MODEL", "deepseek-chat").strip()
    if not api_key:
        raise RuntimeError("缺少 DEEPSEEK_API_KEY，请检查 proj3/.env")
    return LlmConfig(api_key=api_key, base_url=base_url, model=model)


def call_deepseek_json(messages: list[dict[str, str]], cfg: LlmConfig, temperature: float = 0.2) -> dict[str, Any]:
    url = f"{cfg.base_url.rstrip('/')}/v1/chat/completions"
    payload: dict[str, Any] = {
        "model": cfg.model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": 256,
        "response_format": {"type": "json_object"},
    }
    req = request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Authorization": f"Bearer {cfg.api_key}", "Content-Type": "application/json"},
    )
    try:
        with request.urlopen(req, timeout=120) as resp:
            parsed = json.loads(resp.read().decode("utf-8"))
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"DeepSeek HTTP {exc.code}: {detail[:500]}") from exc

    content = parsed.get("choices", [{}])[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError(f"DeepSeek 返回为空: {parsed}")
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"DeepSeek 返回非 JSON: {content[:200]}") from exc


def normalize_numeric_score(result: dict[str, Any]) -> float:
    raw = result.get("score")
    if raw is None:
        raise RuntimeError(f"返回缺少 score 字段: {result}")
    if isinstance(raw, str):
        matched = re.search(r"-?\d+(\.\d+)?", raw)
        if not matched:
            raise RuntimeError(f"score 非数字: {raw}")
        value = float(matched.group(0))
    elif isinstance(raw, (int, float)):
        value = float(raw)
    else:
        raise RuntimeError(f"score 类型错误: {type(raw)}")
    if not (0 <= value <= 5):
        raise RuntimeError(f"score 越界: {value}")
    return value


def get_stage_specific_rules(stage: str) -> str:
    if stage == "post-purchase":
        return (
            "当前为售后反馈（post-purchase）：重点关注实际服务体验、情绪修复、升级人工必要性、品牌与合规风险。"
            "对涉及投诉、误导、过度承诺、隐私担忧、持续负面体验的场景提高风险和紧急评分。"
        )
    return (
        "当前为售前反馈（pre-purchase）：重点关注购买顾虑、信息充分性、转化流失风险和咨询时效。"
        "对涉及隐私疑虑、价格障碍、信任不足、强烈犹豫的场景提高风险和紧急评分。"
    )


def build_dimension_prompt(dimension: str, stage: str, feedback_text: str, customer_profile: str, source_type: str) -> list[dict[str, str]]:
    common = (
        "你是 CRM 优先级标注器，服务于 MindSync（学生与年轻职场人心理陪伴App）。"
        "请只输出 JSON：{\"score\": 0-5数字}，不要输出其它文本。"
        "0表示极低，5表示极高，可用小数。"
    )
    stage_rule = get_stage_specific_rules(stage)
    if dimension == COL_EMOTION:
        rubric = (
            "维度：情绪强度。根据反馈中情绪波动和表达激烈程度评分。"
            "0-1平静客观，2-3有明显情绪，4-5强烈情绪（焦虑、愤怒、崩溃、强烈不满或强烈兴奋）。"
        )
    elif dimension == COL_RISK:
        rubric = (
            "维度：风险等级。根据潜在品牌风险、合规风险、误导风险、舆情扩散风险评分。"
            "0-1基本无风险，2-3存在可控风险，4-5高风险（误导/过度承诺/隐私问题/公开传播可能）。"
        )
    else:
        rubric = (
            "维度：紧急程度。根据响应时效要求评分。"
            "0-1可延后，2-3常规时效处理，4-5需尽快处理（可能流失、升级投诉、快速扩散）。"
        )

    user_content = (
        f"{rubric}\n{stage_rule}\n"
        f"source_type={source_type}\n"
        f"stage={stage}\n"
        f"customer_profile={customer_profile}\n"
        f"feedback_text={feedback_text}\n"
        "仅返回 JSON，如 {\"score\": 3.4}"
    )
    return [{"role": "system", "content": common}, {"role": "user", "content": user_content}]


def score_one_call(
    dimension: str,
    stage: str,
    feedback_text: str,
    customer_profile: str,
    source_type: str,
    cfg: LlmConfig,
) -> float:
    messages = build_dimension_prompt(
        dimension=dimension,
        stage=stage,
        feedback_text=feedback_text,
        customer_profile=customer_profile,
        source_type=source_type,
    )
    result = call_deepseek_json(messages, cfg, temperature=0.2)
    return normalize_numeric_score(result)


def score_with_5_votes(
    dimension: str,
    stage: str,
    feedback_text: str,
    customer_profile: str,
    source_type: str,
    cfg: LlmConfig,
    max_rounds: int = 3,
) -> tuple[float, int]:
    scores: list[float] = []
    failed_calls = 0

    for _ in range(max_rounds):
        needed = 5 - len(scores)
        if needed <= 0:
            break
        with ThreadPoolExecutor(max_workers=needed) as executor:
            futures = [
                executor.submit(
                    score_one_call,
                    dimension,
                    stage,
                    feedback_text,
                    customer_profile,
                    source_type,
                    cfg,
                )
                for _ in range(needed)
            ]
            for fut in as_completed(futures):
                try:
                    scores.append(fut.result())
                except Exception:
                    failed_calls += 1

    if len(scores) < 5:
        raise RuntimeError(
            f"{dimension} 打分失败，stage={stage}，有效分数 {len(scores)}/5，失败调用 {failed_calls}"
        )
    return round(mean(scores), 1), failed_calls


def get_column_index_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(HEADER_ROW, col).value
        if header is None:
            continue
        result[str(header).strip()] = col
    return result


def ensure_target_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    col_map = get_column_index_map(ws)
    next_col = ws.max_column + 1
    for name in TARGET_COLS:
        if name not in col_map:
            ws.cell(HEADER_ROW, next_col).value = name
            col_map[name] = next_col
            next_col += 1
    return col_map


def iter_data_rows(ws: openpyxl.worksheet.worksheet.Worksheet, col_map: dict[str, int]) -> list[int]:
    text_col = col_map.get("feedback_text")
    if not text_col:
        raise RuntimeError("找不到 feedback_text 列")
    rows: list[int] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        text = ws.cell(row, text_col).value
        if text is None:
            continue
        if str(text).strip():
            rows.append(row)
    return rows


def main() -> None:
    cfg = build_llm_config()
    if not XLSX_PATH.exists():
        raise RuntimeError(f"文件不存在: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"工作表不存在: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    col_map = ensure_target_columns(ws)
    rows = iter_data_rows(ws, col_map)
    if not rows:
        raise RuntimeError("未找到可标注的数据行（feedback_text 为空）。")

    stage_col = col_map.get("stage")
    source_col = col_map.get("source_type")
    text_col = col_map.get("feedback_text")
    profile_col = col_map.get("customer_profile")
    if not (stage_col and source_col and text_col and profile_col):
        raise RuntimeError("缺少核心字段列（stage/source_type/feedback_text/customer_profile）")

    retry_failures = 0
    emotion_values: list[float] = []
    risk_values: list[float] = []
    urgency_values: list[float] = []

    for row in rows:
        stage = str(ws.cell(row, stage_col).value or "").strip()
        source_type = str(ws.cell(row, source_col).value or "").strip()
        feedback_text = str(ws.cell(row, text_col).value or "").strip()
        customer_profile = str(ws.cell(row, profile_col).value or "").strip()
        if stage not in {"pre-purchase", "post-purchase"}:
            raise RuntimeError(f"第 {row} 行 stage 非法: {stage}")

        emotion, e_fail = score_with_5_votes(
            dimension=COL_EMOTION,
            stage=stage,
            feedback_text=feedback_text,
            customer_profile=customer_profile,
            source_type=source_type,
            cfg=cfg,
        )
        risk, r_fail = score_with_5_votes(
            dimension=COL_RISK,
            stage=stage,
            feedback_text=feedback_text,
            customer_profile=customer_profile,
            source_type=source_type,
            cfg=cfg,
        )
        urgency, u_fail = score_with_5_votes(
            dimension=COL_URGENCY,
            stage=stage,
            feedback_text=feedback_text,
            customer_profile=customer_profile,
            source_type=source_type,
            cfg=cfg,
        )
        retry_failures += e_fail + r_fail + u_fail

        ws.cell(row, col_map[COL_EMOTION]).value = emotion
        ws.cell(row, col_map[COL_RISK]).value = risk
        ws.cell(row, col_map[COL_URGENCY]).value = urgency

        emotion_values.append(emotion)
        risk_values.append(risk)
        urgency_values.append(urgency)
        print(
            f"[progress] row={row}/{rows[-1]} "
            f"feedback_id={ws.cell(row, col_map.get('feedback_id', 1)).value} "
            f"{COL_EMOTION}={emotion:.1f} {COL_RISK}={risk:.1f} {COL_URGENCY}={urgency:.1f}",
            flush=True,
        )

    wb.save(XLSX_PATH)

    print(f"[done] rows={len(rows)} updated")
    print(f"[columns] {COL_EMOTION}={col_map[COL_EMOTION]}, {COL_RISK}={col_map[COL_RISK]}, {COL_URGENCY}={col_map[COL_URGENCY]}")
    print(f"[retries] failed_calls={retry_failures}")
    print(
        "[ranges] "
        f"emotion={min(emotion_values):.1f}-{max(emotion_values):.1f}, "
        f"risk={min(risk_values):.1f}-{max(risk_values):.1f}, "
        f"urgency={min(urgency_values):.1f}-{max(urgency_values):.1f}"
    )


if __name__ == "__main__":
    main()
