from __future__ import annotations

import json
import os
import random
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib import error, request

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "第7组反馈.xlsx"
SHEET_NAME = "feedback"
ENV_PATH = ROOT / ".env"
HEADER_ROW = 2
DATA_START_ROW = 3
# 与当前表格一致：20 条 human 在第 3–22 行，40 条 agent 回填区为第 23–62 行。
AGENT_BLOCK_START_ROW = 23
AGENT_BLOCK_END_ROW = 62

HEADERS = [
    "feedback_id",
    "source_type",
    "stage",
    "feedback_text",
    "customer_profile",
    "purchase_frequency",
    "avg_order_value",
    "channel",
]
ALLOWED_STAGES = {"pre-purchase", "post-purchase"}
ALLOWED_SOURCE = {"agent"}


@dataclass
class HumanStats:
    pf_min: int
    pf_max: int
    aov_min: float
    aov_max: float
    channels: list[str]
    pre_samples: list[dict[str, Any]]
    post_samples: list[dict[str, Any]]


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def read_human_stats(ws: openpyxl.worksheet.worksheet.Worksheet) -> HumanStats:
    header_cells = [ws.cell(HEADER_ROW, idx + 1).value for idx in range(len(HEADERS))]
    if header_cells != HEADERS:
        raise RuntimeError(
            f"核心字段表头不匹配（第 {HEADER_ROW} 行前 {len(HEADERS)} 列），期望 {HEADERS}，实际 {header_cells}"
        )

    rows: list[dict[str, Any]] = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        values = [ws.cell(row_idx, col_idx + 1).value for col_idx in range(len(HEADERS))]
        if all(v in (None, "") for v in values):
            continue
        record = dict(zip(HEADERS, values))
        rows.append(record)

    human_rows = [r for r in rows if str(r["source_type"]).strip().lower() == "human"]
    if not human_rows:
        raise RuntimeError("未找到 human 样本，无法推断约束范围。")

    purchase_frequency = [int(r["purchase_frequency"]) for r in human_rows]
    avg_order_value = [float(r["avg_order_value"]) for r in human_rows]
    channels = sorted({str(r["channel"]).strip() for r in human_rows if str(r["channel"]).strip()})

    pre_rows = [r for r in human_rows if str(r["stage"]).strip() == "pre-purchase"]
    post_rows = [r for r in human_rows if str(r["stage"]).strip() == "post-purchase"]
    if not pre_rows or not post_rows:
        raise RuntimeError("human 样本中缺少售前或售后数据，无法做分阶段 few-shot。")

    random.seed(7)
    pre_samples = random.sample(pre_rows, k=min(4, len(pre_rows)))
    post_samples = random.sample(post_rows, k=min(4, len(post_rows)))

    return HumanStats(
        pf_min=min(purchase_frequency),
        pf_max=max(purchase_frequency),
        aov_min=min(avg_order_value),
        aov_max=max(avg_order_value),
        channels=channels,
        pre_samples=pre_samples,
        post_samples=post_samples,
    )


def normalize_channel_text(value: str) -> str:
    text = re.sub(r"\s+", "", value.strip()).replace("Ａ", "A").replace("ａ", "a")
    text = text.replace("APP", "app")
    return text


def build_common_context(stats: HumanStats) -> str:
    return (
        "你在生成 MindSync 的结构化消费者反馈数据。"
        "MindSync 是面向学生与初入职场白领的对话式 AI 心理陪伴与情绪管理 App，"
        "核心功能为聊天、心理学建模、回顾与建议。\n"
        "目标用户：学生与年轻职场人。\n"
        "必须严格输出 JSON 对象，格式为 {\"records\":[...]}。\n"
        "每条记录只允许包含以下 8 个字段："
        "feedback_id, source_type, stage, feedback_text, customer_profile, purchase_frequency, avg_order_value, channel。\n"
        "字段规则：\n"
        f"- source_type 必须为 'agent'\n"
        f"- stage 必须为 'pre-purchase' 或 'post-purchase'\n"
        f"- purchase_frequency 必须为整数，范围 {stats.pf_min} 到 {stats.pf_max}\n"
        f"- avg_order_value 必须为数值，范围 {stats.aov_min:.0f} 到 {stats.aov_max:.0f}\n"
        f"- channel 必须从以下集合选取：{', '.join(stats.channels)}\n"
        "- feedback_text 必须是自然中文，避免和示例重复或高度相似\n"
        "- customer_profile 需包含年龄段、场景、动机，且与 feedback_text 自洽\n"
        "- 不要输出 markdown，不要输出解释文字，只输出 JSON"
    )


def build_fewshot(samples: list[dict[str, Any]]) -> str:
    rows = []
    for item in samples:
        rows.append(
            {
                "feedback_id": item["feedback_id"],
                "source_type": item["source_type"],
                "stage": item["stage"],
                "feedback_text": item["feedback_text"],
                "customer_profile": item["customer_profile"],
                "purchase_frequency": item["purchase_frequency"],
                "avg_order_value": item["avg_order_value"],
                "channel": item["channel"],
            }
        )
    return json.dumps({"examples": rows}, ensure_ascii=False, indent=2)


def sentiment_instruction(stage: str) -> str:
    if stage == "pre-purchase":
        return (
            "情绪分布要求（20条）:"
            "正向约10条、中性约6条、负向约4条。"
            "负向内容可以涉及价格顾虑、隐私担忧、功能疑问、担心效果不稳定。"
            "负向样本的 feedback_text 里请明确出现负向表达词（如：担心/不放心/犹豫/太贵/失望/不满）。"
        )
    return (
        "情绪分布要求（20条）:"
        "正向约6条、中性约6条、负向约8条。"
        "负向内容应覆盖真实售后问题，例如回复太模板化、情绪失配、体验卡顿、建议不准确、需要人工升级等。"
        "负向样本的 feedback_text 里请明确出现负向表达词（如：失望/不满/敷衍/冷冰冰/卡顿/退款/不推荐）。"
    )


def call_deepseek_json(messages: list[dict[str, str]], model: str, base_url: str, api_key: str) -> dict[str, Any]:
    url = f"{base_url.rstrip('/')}/v1/chat/completions"
    payload: dict[str, Any] = {
        "model": model,
        "messages": messages,
        "temperature": 0.9,
        "max_tokens": 4096,
        "response_format": {"type": "json_object"},
    }
    req = request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
    )
    try:
        with request.urlopen(req, timeout=180) as resp:
            parsed = json.loads(resp.read().decode("utf-8"))
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"DeepSeek HTTP {exc.code}: {detail[:500]}") from exc

    content = parsed.get("choices", [{}])[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError(f"DeepSeek 返回为空: {parsed}")
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"DeepSeek 返回非 JSON: {content[:500]}") from exc


def normalize_record(item: dict[str, Any], stage: str, row_num: int, stats: HumanStats) -> dict[str, Any]:
    missing = [k for k in HEADERS if k not in item]
    extra = [k for k in item if k not in HEADERS]
    if missing or extra:
        raise RuntimeError(f"记录#{row_num} 字段异常，缺失={missing}，多余={extra}")

    source_type = str(item["source_type"]).strip()
    stage_value = str(item["stage"]).strip()
    if source_type not in ALLOWED_SOURCE:
        raise RuntimeError(f"记录#{row_num} source_type 非法: {source_type}")
    if stage_value != stage:
        raise RuntimeError(f"记录#{row_num} stage 应为 {stage}，实际 {stage_value}")
    if stage_value not in ALLOWED_STAGES:
        raise RuntimeError(f"记录#{row_num} stage 非法: {stage_value}")

    feedback_text = re.sub(r"\s+", " ", str(item["feedback_text"]).strip())
    customer_profile = re.sub(r"\s+", " ", str(item["customer_profile"]).strip())
    if len(feedback_text) < 8:
        raise RuntimeError(f"记录#{row_num} feedback_text 过短")
    if len(customer_profile) < 8:
        raise RuntimeError(f"记录#{row_num} customer_profile 过短")

    try:
        purchase_frequency = int(item["purchase_frequency"])
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"记录#{row_num} purchase_frequency 非整数") from exc
    if not (stats.pf_min <= purchase_frequency <= stats.pf_max):
        raise RuntimeError(
            f"记录#{row_num} purchase_frequency 越界: {purchase_frequency}, 期望 {stats.pf_min}-{stats.pf_max}"
        )

    try:
        avg_order_value = float(item["avg_order_value"])
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"记录#{row_num} avg_order_value 非数字") from exc
    if not (stats.aov_min <= avg_order_value <= stats.aov_max):
        raise RuntimeError(
            f"记录#{row_num} avg_order_value 越界: {avg_order_value}, 期望 {stats.aov_min}-{stats.aov_max}"
        )

    channel = str(item["channel"]).strip()
    channel_map = {normalize_channel_text(c): c for c in stats.channels}
    normalized_key = normalize_channel_text(channel)
    if normalized_key not in channel_map:
        raise RuntimeError(f"记录#{row_num} channel 不在候选集合: {channel}")
    channel = channel_map[normalized_key]

    return {
        "feedback_id": "",
        "source_type": source_type,
        "stage": stage_value,
        "feedback_text": feedback_text,
        "customer_profile": customer_profile,
        "purchase_frequency": purchase_frequency,
        "avg_order_value": round(avg_order_value, 1),
        "channel": channel,
    }


def generate_stage_records(stage: str, count: int, stats: HumanStats, model: str, base_url: str, api_key: str) -> list[dict[str, Any]]:
    samples = stats.pre_samples if stage == "pre-purchase" else stats.post_samples
    system_prompt = build_common_context(stats)
    base_user_prompt = (
        f"任务：生成 {count} 条 stage={stage} 的反馈记录。\n"
        "注意：feedback_id 先留空字符串，后续由脚本统一赋值。\n"
        f"{sentiment_instruction(stage)}\n"
        f"few-shot 参考（不可复写）:\n{build_fewshot(samples)}"
    )

    last_error = ""
    for attempt in range(1, 4):
        user_prompt = base_user_prompt
        if last_error:
            user_prompt += f"\n上一次结果问题：{last_error}\n请修正后重新生成完整20条。"
        result = call_deepseek_json(
            messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
            model=model,
            base_url=base_url,
            api_key=api_key,
        )
        records = result.get("records")
        if not isinstance(records, list) or len(records) != count:
            last_error = (
                f"{stage} 生成条数不正确，期望 {count}，实际 {len(records) if isinstance(records, list) else records}"
            )
            continue

        normalized: list[dict[str, Any]] = []
        try:
            for idx, item in enumerate(records, start=1):
                if not isinstance(item, dict):
                    raise RuntimeError(f"{stage} 记录#{idx} 不是对象")
                normalized.append(normalize_record(item, stage=stage, row_num=idx, stats=stats))

            text_set = set()
            for idx, item in enumerate(normalized, start=1):
                key = item["feedback_text"]
                if key in text_set:
                    raise RuntimeError(f"{stage} 出现重复 feedback_text，记录#{idx}")
                text_set.add(key)

            return normalized
        except RuntimeError as exc:
            last_error = str(exc)
            if attempt == 3:
                raise
            continue

    raise RuntimeError(f"{stage} 生成失败: {last_error}")


def clear_write_region(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, end_row: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(r, c).value = None


def write_records(ws: openpyxl.worksheet.worksheet.Worksheet, records: list[dict[str, Any]], start_row: int, id_start: int) -> None:
    for i, record in enumerate(records):
        row = start_row + i
        record = dict(record)
        record["feedback_id"] = f"G07-{id_start + i:03d}"
        values = [record[h] for h in HEADERS]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col).value = value


def main() -> None:
    load_env_file(ENV_PATH)
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com").strip()
    model = os.getenv("DEEPSEEK_MODEL", "deepseek-chat").strip()
    if not api_key:
        raise RuntimeError("缺少 DEEPSEEK_API_KEY，请在 proj3/.env 中配置。")
    if not XLSX_PATH.exists():
        raise RuntimeError(f"文件不存在: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"工作表不存在: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    stats = read_human_stats(ws)
    print(
        f"[human_stats] purchase_frequency={stats.pf_min}-{stats.pf_max}, "
        f"avg_order_value={stats.aov_min:.0f}-{stats.aov_max:.0f}, channels={len(stats.channels)}"
    )

    pre_records = generate_stage_records("pre-purchase", 20, stats, model, base_url, api_key)
    post_records = generate_stage_records("post-purchase", 20, stats, model, base_url, api_key)
    clear_write_region(ws, start_row=AGENT_BLOCK_START_ROW, end_row=AGENT_BLOCK_END_ROW)
    write_records(ws, pre_records, start_row=AGENT_BLOCK_START_ROW, id_start=21)
    write_records(ws, post_records, start_row=AGENT_BLOCK_START_ROW + 20, id_start=41)

    wb.save(XLSX_PATH)
    print(f"[write_back] 覆盖写入完成: rows {AGENT_BLOCK_START_ROW}-{AGENT_BLOCK_END_ROW}")
    print("[summary] pre-purchase=20, post-purchase=20, ids=G07-021..G07-060")


if __name__ == "__main__":
    main()
