from __future__ import annotations

import json
import os
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib import error, request

import openpyxl

ROOT = Path(__file__).resolve().parent
ENV_PATH = ROOT / ".env"
XLSX_PATH = ROOT / "第7组反馈.xlsx"
SHEET_NAME = "feedback"
# 与当前 xlsx 一致：第 1 行为分组表头，第 2 行为字段名，数据从第 3 行起。
HEADER_ROW = 2
DATA_START_ROW = 3

# 写入表的列（与表中「标准输出」区字段名一致）；「优先级评分」由脚本留空，供队友在 3.4 中填写。
OUTPUT_COLUMNS = [
    "Agent回复",
    "问题归类",
    "情绪判断",
    "客户价值层级",
    "建议处理方式",
    "建议人工介入",
    "人工介入理由",
    "优先级评分",
]

REQUIRED_COLUMNS = [
    "feedback_id",
    "source_type",
    "stage",
    "feedback_text",
    "customer_profile",
    "purchase_frequency",
    "avg_order_value",
]

ROADMAP_KNOWLEDGE = (
    "MindSync 产品迭代路径（用于售后承诺边界）：\n"
    "1) 在迭代中：对话稳定性、情绪识别准确度、陪伴连续性、个性化建议质量、APP性能卡顿优化。\n"
    "2) 暂不承诺：替代线下心理治疗、替用户操作手机、面向老年人专门陪伴功能、医疗诊断结论。\n"
    "3) 沟通原则：感谢反馈、不夸大承诺、给出可执行下一步、保持温和可靠语气。"
)

# 名义上的「RAG」：仅将话术以 topic : answer 形式写入 system，不做检索与相似度；动态人类话术回灌未实装。
NOMINAL_KB = [
    {
        "topic": "隐私与数据安全",
        "answer": "MindSync 的所有用户聊天文本和画像数据都保存在用户的设备上，服务器不存储相关内容。",
    },
    {
        "topic": "个性化与推荐逻辑",
        "answer": "MindSync 会基于用户对话历史和偏好进行多维度心理建模，提供个性化心理回复。",
    },
    {
        "topic": "价格与体验",
        "answer": "MindSync 提供3天免费试用，用户可先体验核心功能再决定是否升级付费方案。",
    },
    {
        "topic": "人工支持与问题反馈",
        "answer": "遇到复杂或个性化问题时可由人工客服进一步跟进，确保反馈被完整记录和处理。",
    },
]


@dataclass
class LlmConfig:
    api_key: str
    base_url: str
    model: str


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def build_llm_config() -> LlmConfig:
    load_env_file(ENV_PATH)
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com").strip()
    model = os.getenv("DEEPSEEK_MODEL", "deepseek-chat").strip()
    if not api_key:
        raise RuntimeError("缺少 DEEPSEEK_API_KEY，请在 proj3/.env 中配置。")
    return LlmConfig(api_key=api_key, base_url=base_url, model=model)


def call_deepseek_json(messages: list[dict[str, str]], cfg: LlmConfig, temperature: float = 0.4) -> dict[str, Any]:
    url = f"{cfg.base_url.rstrip('/')}/v1/chat/completions"
    payload: dict[str, Any] = {
        "model": cfg.model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": 2048,
        "response_format": {"type": "json_object"},
    }
    req = request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Authorization": f"Bearer {cfg.api_key}", "Content-Type": "application/json"},
    )
    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            with request.urlopen(req, timeout=180) as resp:
                parsed = json.loads(resp.read().decode("utf-8"))
                break
        except error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            raise RuntimeError(f"DeepSeek HTTP {exc.code}: {detail[:500]}") from exc
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if attempt == 3:
                raise RuntimeError(f"DeepSeek 请求失败(重试3次): {exc}") from exc
            time.sleep(2 * attempt)
    else:
        raise RuntimeError(f"DeepSeek 请求失败: {last_error}")

    content = parsed.get("choices", [{}])[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError(f"DeepSeek 返回为空: {parsed}")
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"DeepSeek 返回非 JSON: {content[:500]}") from exc


def nominal_kb_block() -> str:
    lines = [f"{item['topic']} : {item['answer']}" for item in NOMINAL_KB]
    return (
        "以下为名义上的 RAG 知识库（以 topic : answer 形式提供，非向量检索；"
        "人类话术动态回灌知识库未实装）。\n"
        + "\n".join(lines)
    )


def build_generator_system_prompt() -> str:
    return (
        "你是 MindSync 的 CRM 回复助手。你要输出结构化 JSON，语气温和、可信、不过度承诺。"
        "重点遵循：售前强调解释与体验引导；售后强调情绪修复、边界清晰和可执行建议。"
        "严禁虚构产品能力，严禁医疗化承诺。\n"
        "撰写回复时可自行把握语气轻重与紧迫感的分寸，无需在 JSON 中输出任何数值型「优先级评分」字段；"
        "也不要引用或复述表格中可能存在的「情绪强度」「风险等级」「紧急程度」等列的数值。\n\n"
        f"{ROADMAP_KNOWLEDGE}\n\n"
        f"{nominal_kb_block()}"
    )


def parse_float(value: Any, name: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise RuntimeError(f"{name} 非数值: {value}") from exc


def value_tier_by_norm(norm_value: float) -> str:
    if norm_value >= 1.2:
        return "高价值"
    if norm_value <= 0.8:
        return "低价值"
    return "中价值"


def build_generator_messages(
    *,
    stage: str,
    feedback_text: str,
    customer_profile: str,
    value_tier: str,
    retry_feedback: str = "",
) -> list[dict[str, str]]:
    system_prompt = build_generator_system_prompt()
    if stage == "pre-purchase":
        stage_policy = (
            "售前策略：问题通常由“信息不清楚”+“潜在购买意图”组成。"
            "先回答关键信息，再在自然语气下邀请体验（可弱化，不可强推）。"
        )
    else:
        stage_policy = (
            "售后策略：先判断好评/明确不满/不清晰差评。"
            "好评要个性化感谢；明确不满要按迭代边界回应；不清晰差评先礼貌追问细节。"
        )

    user_prompt = (
        f"{stage_policy}\n"
        "请仅输出 JSON，字段必须完整："
        "{\"问题归类\":\"...\",\"情绪判断\":\"...\",\"建议处理方式\":\"...\","
        "\"是否建议人工介入\":\"是/否\",\"人工介入理由\":\"...\",\"回复草案\":\"...\"}\n"
        "其中「是否建议人工介入」「人工介入理由」须由你根据本条反馈综合判断一次性给出，不要依赖任何外部硬规则标签。\n"
        f"输入：stage={stage}\n"
        f"customer_profile={customer_profile}\n"
        f"feedback_text={feedback_text}\n"
        f"客户价值层级（已由脚本根据 purchase_frequency×avg_order_value 相对全表均值归一化后分档，仅作参考）={value_tier}\n"
    )
    if retry_feedback:
        user_prompt += f"上轮验收未通过原因：{retry_feedback}\n请据此重写并改进语言风格。"
    return [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}]


def build_acceptance_messages(reply_text: str, stage: str) -> list[dict[str, str]]:
    system_prompt = (
        "你是话术验收助手，只检查语言风格与风险，不重写。"
        "请仅输出 JSON：{\"pass\":true/false,\"feedback\":\"...\"}。"
        "若通过，feedback 写“通过”。若不通过，给出1-3条可执行修改意见。"
    )
    user_prompt = (
        f"stage={stage}\n"
        "验收维度：语气温和可信、不过度承诺、避免AI味和讨好感、与心理陪伴品牌调性一致。\n"
        f"待验收回复：{reply_text}"
    )
    return [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}]


def normalize_yes_no(value: str) -> str:
    text = str(value).strip()
    if text in {"是", "需要", "建议"}:
        return "是"
    if text in {"否", "不需要", "无需"}:
        return "否"
    if "不" in text and "人工" in text:
        return "否"
    if "人工" in text or "升级" in text:
        return "是"
    return "否"


def intervention_from_llm(raw_yes: str, raw_reason: str) -> tuple[str, str]:
    yn = normalize_yes_no(raw_yes)
    reason = str(raw_reason).strip()
    if yn == "是":
        return "是", reason or "综合反馈内容，建议人工跟进。"
    return "否", reason or "可由 Agent 直接承接。"


def get_column_index_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(HEADER_ROW, col).value
        if header is None:
            continue
        col_map[str(header).strip()] = col
    return col_map


def ensure_output_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    col_map = get_column_index_map(ws)
    next_col = ws.max_column + 1
    for name in OUTPUT_COLUMNS:
        if name not in col_map:
            ws.cell(HEADER_ROW, next_col).value = name
            col_map[name] = next_col
            next_col += 1
    return col_map


def validate_required_columns(col_map: dict[str, int]) -> None:
    missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing:
        raise RuntimeError(f"缺少必要列: {missing}")


def iter_data_rows(ws: openpyxl.worksheet.worksheet.Worksheet, text_col: int) -> list[int]:
    rows: list[int] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        value = ws.cell(row, text_col).value
        if value is None:
            continue
        if str(value).strip():
            rows.append(row)
    return rows


def calc_value_mean(ws: openpyxl.worksheet.worksheet.Worksheet, rows: list[int], col_map: dict[str, int]) -> float:
    values: list[float] = []
    for row in rows:
        pf = parse_float(ws.cell(row, col_map["purchase_frequency"]).value, "purchase_frequency")
        aov = parse_float(ws.cell(row, col_map["avg_order_value"]).value, "avg_order_value")
        values.append(pf * aov)
    if not values:
        raise RuntimeError("没有可计算客户价值的数据行。")
    return sum(values) / len(values)


def generate_one_row(
    *,
    stage: str,
    feedback_text: str,
    customer_profile: str,
    value_tier: str,
    cfg: LlmConfig,
) -> dict[str, str]:
    last_feedback = ""
    last_data: dict[str, Any] | None = None
    for _ in range(5):
        messages = build_generator_messages(
            stage=stage,
            feedback_text=feedback_text,
            customer_profile=customer_profile,
            value_tier=value_tier,
            retry_feedback=last_feedback,
        )
        data = call_deepseek_json(messages, cfg, temperature=0.45)
        last_data = data
        reply = str(data.get("回复草案", "")).strip()
        if not reply:
            last_feedback = "回复草案为空，请补全。"
            continue
        review = call_deepseek_json(build_acceptance_messages(reply, stage), cfg, temperature=0.1)
        passed = bool(review.get("pass"))
        feedback = str(review.get("feedback", "")).strip()
        if passed:
            intervention, reason = intervention_from_llm(
                str(data.get("是否建议人工介入", "")),
                str(data.get("人工介入理由", "")),
            )
            return {
                "Agent回复": reply,
                "问题归类": str(data.get("问题归类", "其他")).strip() or "其他",
                "情绪判断": str(data.get("情绪判断", "中性")).strip() or "中性",
                "建议处理方式": str(data.get("建议处理方式", "标准回复")).strip() or "标准回复",
                "建议人工介入": intervention,
                "人工介入理由": reason,
            }
        last_feedback = feedback if feedback else "语气或风格未通过，请重写。"

    fallback = last_data or {}
    intervention, reason = intervention_from_llm(
        str(fallback.get("是否建议人工介入", "")),
        str(fallback.get("人工介入理由", "")),
    )
    return {
        "Agent回复": str(fallback.get("回复草案", "感谢你的反馈，我们会继续优化服务。")).strip()
        or "感谢你的反馈，我们会继续优化服务。",
        "问题归类": str(fallback.get("问题归类", "其他")).strip() or "其他",
        "情绪判断": str(fallback.get("情绪判断", "中性")).strip() or "中性",
        "建议处理方式": str(fallback.get("建议处理方式", "标准回复")).strip() or "标准回复",
        "建议人工介入": intervention,
        "人工介入理由": reason,
    }


def main() -> None:
    cfg = build_llm_config()
    if not XLSX_PATH.exists():
        raise RuntimeError(f"文件不存在: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"工作表不存在: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    col_map = get_column_index_map(ws)
    validate_required_columns(col_map)
    text_col = col_map["feedback_text"]
    rows = iter_data_rows(ws, text_col)
    if not rows:
        raise RuntimeError("没有可处理的反馈行。")

    value_mean = calc_value_mean(ws, rows, col_map)
    col_map = ensure_output_columns(ws)
    priority_col = col_map.get("优先级评分")

    for row in rows:
        feedback_id = str(ws.cell(row, col_map["feedback_id"]).value or "").strip()
        stage = str(ws.cell(row, col_map["stage"]).value or "").strip()
        feedback_text = str(ws.cell(row, col_map["feedback_text"]).value or "").strip()
        customer_profile = str(ws.cell(row, col_map["customer_profile"]).value or "").strip()
        pf = parse_float(ws.cell(row, col_map["purchase_frequency"]).value, "purchase_frequency")
        aov = parse_float(ws.cell(row, col_map["avg_order_value"]).value, "avg_order_value")

        value_raw = pf * aov
        value_norm = value_raw / value_mean if value_mean else 1.0
        value_tier = value_tier_by_norm(value_norm)

        result = generate_one_row(
            stage=stage,
            feedback_text=feedback_text,
            customer_profile=customer_profile,
            value_tier=value_tier,
            cfg=cfg,
        )
        result["客户价值层级"] = value_tier

        for column in OUTPUT_COLUMNS:
            if column == "优先级评分":
                if priority_col:
                    ws.cell(row, priority_col).value = None
                continue
            ws.cell(row, col_map[column]).value = result[column]

        print(
            f"[progress] row={row}/{rows[-1]} id={feedback_id} stage={stage} value={value_tier}",
            flush=True,
        )

    wb.save(XLSX_PATH)
    print(f"[done] rows={len(rows)} updated, saved={XLSX_PATH.name}, cols={OUTPUT_COLUMNS} (优先级评分已留空)")


if __name__ == "__main__":
    main()
