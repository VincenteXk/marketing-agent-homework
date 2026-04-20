from __future__ import annotations

import argparse
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib import error, request

import openpyxl

ROOT = Path(__file__).resolve().parent
ENV_PATH = ROOT / ".env"
XLSX_PATH = ROOT / "第7组反馈.xlsx"
SHEET_NAME = "feedback"
HEADER_ROW = 2
DATA_START_ROW = 3

COL_HUMAN_TEXT = "人工回复内容"
COL_AGENT_TEXT = "Agent回复"

# 同一套评价体系分别盲评两条回复；列顺序：每维先人工、后 Agent（表头含来源，提示词不含来源）
DIMENSION_BASES = ["情感共鸣度", "准确性与一致性", "潜在品牌风险", "可扩展性"]

OUTPUT_COLS: list[str] = []
for base in DIMENSION_BASES:
    OUTPUT_COLS.append(f"{base}-人工")
    OUTPUT_COLS.append(f"{base}-Agent")

MAX_ATTEMPTS_PER_DIM = 3

MINDSYNC_BRAND = (
    "MindSync：面向学生与年轻职场人的对话式 AI 心理陪伴与情绪管理 App；"
    "语气应温和、可信、不过度承诺。"
    "边界：不可替代线下心理治疗、不作医疗诊断、不替用户操作手机；"
    "数据与隐私表述须谨慎，避免夸大疗效或功能。"
)

# 评分语义：score 为 0–5，表示「该条待评客服回复」在该维度上的程度，越高越好。
# 「潜在品牌风险」维度：5 表示表述安全、品牌与合规风险很低；0 表示风险很高。


@dataclass
class LlmConfig:
    api_key: str
    base_url: str
    model: str


@dataclass(frozen=True)
class DimensionSpec:
    """单一评价维度（与人工/Agent 无关，用于两套完全相同的盲评提示）。"""
    column_name: str
    system_prompt: str


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def build_llm_config() -> LlmConfig:
    load_env_file(ENV_PATH)
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    base_url = os.getenv("DEEPSEEK_BASE_URL", "https://api.deepseek.com").strip()
    model = os.getenv("DEEPSEEK_MODEL", "deepseek-chat").strip()
    if not api_key:
        raise RuntimeError("缺少 DEEPSEEK_API_KEY，请检查 proj3/.env")
    return LlmConfig(api_key=api_key, base_url=base_url, model=model)


def call_deepseek_json(messages: list[dict[str, str]], cfg: LlmConfig, temperature: float = 0.25) -> dict[str, Any]:
    url = f"{cfg.base_url.rstrip('/')}/v1/chat/completions"
    payload: dict[str, Any] = {
        "model": cfg.model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": 768,
        "response_format": {"type": "json_object"},
    }
    req = request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={"Authorization": f"Bearer {cfg.api_key}", "Content-Type": "application/json"},
    )
    try:
        with request.urlopen(req, timeout=120) as resp:
            parsed = json.loads(resp.read().decode("utf-8"))
    except error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"DeepSeek HTTP {exc.code}: {detail[:500]}") from exc

    content = parsed.get("choices", [{}])[0].get("message", {}).get("content", "")
    if not content:
        raise RuntimeError(f"DeepSeek 返回为空: {parsed}")
    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"DeepSeek 返回非 JSON: {content[:200]}") from exc


def parse_score(raw: Any) -> float:
    if raw is None:
        raise RuntimeError("缺少 score")
    if isinstance(raw, str):
        matched = re.search(r"-?\d+(\.\d+)?", raw)
        if not matched:
            raise RuntimeError(f"score 非数字: {raw}")
        value = float(matched.group(0))
    elif isinstance(raw, (int, float)):
        value = float(raw)
    else:
        raise RuntimeError(f"score 类型错误: {type(raw)}")
    if not (0 <= value <= 5):
        raise RuntimeError(f"score 越界: {value}")
    return value


def format_score_for_cell(score: float) -> str:
    rounded = round(score, 1)
    if rounded == int(rounded):
        return str(int(rounded))
    return f"{rounded:.1f}".rstrip("0").rstrip(".")


def normalize_rationale(text: str) -> str:
    return text.strip().replace("\n", " ")


def build_dimension_specs() -> list[DimensionSpec]:
    blind_rules = (
        f"\n{MINDSYNC_BRAND}\n\n"
        "你将看到「客户反馈与情境」以及单独一条「待评价的客服回复文本」。\n"
        "严禁推断或说明该回复来自人工还是自动化；只依据文本本身按本维度量表打分。\n"
        "rationale 仅写本条回复的给分依据（简洁完整即可），不要与其它回复对比。\n"
        "仅输出 JSON：{\"score\": 0到5的数字（可用一位小数）, \"rationale\": \"...\"}。\n"
        "不要输出其它键或解释。"
    )
    return [
        DimensionSpec(
            column_name="情感共鸣度",
            system_prompt=(
                "你是 CRM 质量评审员，维度：情感共鸣度。"
                "0–5 表示该客服回复对客户情绪与处境的理解与共情是否到位；"
                "5 表示共情自然、称呼与语气贴切；0 表示冷漠、脱节或敷衍。"
                + blind_rules
            ),
        ),
        DimensionSpec(
            column_name="准确性与一致性",
            system_prompt=(
                "你是 CRM 质量评审员，维度：准确性与一致性（信息准确 + 品牌口径一致）。"
                "0–5 表示该客服回复是否与客户问题及 MindSync 产品事实、承诺边界一致；"
                "5 表示信息准确且口径统一；0 表示明显误导、夸大或与品牌调性冲突。"
                + blind_rules
            ),
        ),
        DimensionSpec(
            column_name="潜在品牌风险",
            system_prompt=(
                "你是 CRM 合规与品牌风险评审员，维度：潜在品牌风险（分数越高越安全）。"
                "0–5 表示该客服回复在过度承诺、医疗化表述、隐私误导、激化矛盾等方面的安全程度；"
                "5 表示几乎无品牌与合规风险；0 表示风险很高。"
                + blind_rules
            ),
        ),
        DimensionSpec(
            column_name="可扩展性",
            system_prompt=(
                "你是 CRM 运营评审员，维度：可扩展性（模板化潜力）。"
                "0–5 表示在保持得体前提下，该客服回复是否易于沉淀为可复用话术/模板；"
                "5 表示结构清晰、变量少、易批量复用；0 表示过于个案化或强依赖上下文难以模板化。"
                + blind_rules
            ),
        ),
    ]


def build_blind_user_payload(
    *,
    stage: str,
    source_type: str,
    feedback_text: str,
    customer_profile: str,
    dimension_label: str,
    reply_text: str,
) -> str:
    return (
        f"评价维度：{dimension_label}\n"
        f"客户所处阶段 stage={stage}\n"
        f"source_type={source_type}\n"
        f"customer_profile={customer_profile}\n"
        f"客户反馈原文 feedback_text={feedback_text}\n\n"
        f"待评价的客服回复文本=\n{reply_text}\n"
    )


def score_one_reply(
    spec: DimensionSpec,
    *,
    output_col: str,
    stage: str,
    source_type: str,
    feedback_text: str,
    customer_profile: str,
    reply_text: str,
    cfg: LlmConfig,
) -> str:
    messages = [
        {"role": "system", "content": spec.system_prompt},
        {
            "role": "user",
            "content": build_blind_user_payload(
                stage=stage,
                source_type=source_type,
                feedback_text=feedback_text,
                customer_profile=customer_profile,
                dimension_label=spec.column_name,
                reply_text=reply_text,
            ),
        },
    ]
    for attempt in range(1, MAX_ATTEMPTS_PER_DIM + 1):
        try:
            result = call_deepseek_json(messages, cfg, temperature=0.25)
            score = parse_score(result.get("score"))
            rationale = normalize_rationale(str(result.get("rationale", "")))
            if not rationale:
                rationale = "未说明"
            score_s = format_score_for_cell(score)
            return f"{score_s}分，{rationale}"
        except Exception as exc:  # noqa: BLE001
            print(f"[retry] {output_col} attempt {attempt}/{MAX_ATTEMPTS_PER_DIM}: {exc}", flush=True)
    return f"错误：评分失败（{output_col}）"


def get_column_index_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(HEADER_ROW, col).value
        if header is None:
            continue
        result[str(header).strip()] = col
    return result


def ensure_output_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    """在表最右侧依次追加 OUTPUT_COLS 中缺失的列（第 2 行写表头）。"""
    col_map = get_column_index_map(ws)
    next_col = ws.max_column + 1
    for name in OUTPUT_COLS:
        if name not in col_map:
            ws.cell(HEADER_ROW, next_col).value = name
            col_map[name] = next_col
            next_col += 1
    return col_map


def parse_force_rows(arg: str | None) -> set[int] | None:
    if not arg or not str(arg).strip():
        return None
    out: set[int] = set()
    for part in str(arg).split(","):
        part = part.strip()
        if not part:
            continue
        out.add(int(part))
    return out or None


def rows_to_fill_none(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    text_col: int,
    scored_rows: set[int],
    force_rows: set[int] | None,
) -> list[int]:
    """有 feedback 文本、且本次未做 LLM 评分的行，八列填「无」。"""
    out: list[int] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if force_rows is not None and row not in force_rows:
            continue
        if not str(ws.cell(row, text_col).value or "").strip():
            continue
        if row in scored_rows:
            continue
        out.append(row)
    return out


def write_none_for_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col_map: dict[str, int]) -> None:
    for name in OUTPUT_COLS:
        ws.cell(row, col_map[name]).value = "无"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="2.4：同一套四维标准分别盲评人工与 Agent 回复，写入表末八列。"
    )
    parser.add_argument(
        "--force-rows",
        type=str,
        default="",
        help="仅处理指定 Excel 行号（逗号分隔，如 3,5,12）；仍要求该行「人工回复内容」与「Agent回复」均非空。",
    )
    args = parser.parse_args()
    force_rows = parse_force_rows(args.force_rows)

    if not XLSX_PATH.exists():
        raise RuntimeError(f"文件不存在: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"工作表不存在: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    col_map = get_column_index_map(ws)
    for required in (COL_HUMAN_TEXT, COL_AGENT_TEXT, "feedback_text", "stage", "customer_profile"):
        if required not in col_map:
            raise RuntimeError(f"缺少必要列: {required}")

    human_col = col_map[COL_HUMAN_TEXT]
    agent_col = col_map[COL_AGENT_TEXT]
    text_col = col_map["feedback_text"]
    stage_col = col_map["stage"]
    profile_col = col_map["customer_profile"]
    source_col = col_map.get("source_type")

    target_rows: list[int] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if force_rows is not None and row not in force_rows:
            continue
        h = ws.cell(row, human_col).value
        a = ws.cell(row, agent_col).value
        if not (str(h or "").strip() and str(a or "").strip()):
            continue
        if not str(ws.cell(row, text_col).value or "").strip():
            continue
        target_rows.append(row)

    col_map = ensure_output_columns(ws)
    scored_set = set(target_rows)

    if not target_rows:
        print(
            "[info] 没有同时含「人工回复内容」与「Agent回复」的可评分行；"
            "仍会为范围内非输出行的八列写入「无」。",
            flush=True,
        )
    else:
        print(f"[info] 待评分行数={len(target_rows)} rows={target_rows[:5]}{'...' if len(target_rows) > 5 else ''}", flush=True)

    cfg: LlmConfig | None = None
    if target_rows:
        cfg = build_llm_config()
    specs = build_dimension_specs()
    if [s.column_name for s in specs] != DIMENSION_BASES:
        raise RuntimeError("维度定义与 DIMENSION_BASES 不一致")

    for row in target_rows:
        assert cfg is not None
        stage = str(ws.cell(row, stage_col).value or "").strip()
        feedback_text = str(ws.cell(row, text_col).value or "").strip()
        customer_profile = str(ws.cell(row, profile_col).value or "").strip()
        human_reply = str(ws.cell(row, human_col).value or "").strip()
        agent_reply = str(ws.cell(row, agent_col).value or "").strip()
        source_type = str(ws.cell(row, source_col).value or "").strip() if source_col else ""

        def run_one(output_col: str, spec: DimensionSpec, reply: str) -> tuple[str, str]:
            cell = score_one_reply(
                spec,
                output_col=output_col,
                stage=stage,
                source_type=source_type,
                feedback_text=feedback_text,
                customer_profile=customer_profile,
                reply_text=reply,
                cfg=cfg,
            )
            return output_col, cell

        results: dict[str, str] = {}
        with ThreadPoolExecutor(max_workers=8) as executor:
            futures = []
            for spec in specs:
                futures.append(
                    executor.submit(
                        run_one,
                        f"{spec.column_name}-人工",
                        spec,
                        human_reply,
                    )
                )
                futures.append(
                    executor.submit(
                        run_one,
                        f"{spec.column_name}-Agent",
                        spec,
                        agent_reply,
                    )
                )
            for fut in as_completed(futures):
                col_name, cell_val = fut.result()
                results[col_name] = cell_val

        for name in OUTPUT_COLS:
            ws.cell(row, col_map[name]).value = results[name]

        fid = ws.cell(row, col_map.get("feedback_id", 1)).value
        print(f"[progress] row={row} feedback_id={fid} ok", flush=True)

    none_rows = rows_to_fill_none(ws, text_col=text_col, scored_rows=scored_set, force_rows=force_rows)
    for row in none_rows:
        write_none_for_row(ws, row, col_map)
    if none_rows:
        print(f"[info] 非输出行已填「无」: count={len(none_rows)}", flush=True)

    wb.save(XLSX_PATH)
    print(
        f"[done] 盲评写入 {len(target_rows)} 行；非输出填「无」 {len(none_rows)} 行；saved={XLSX_PATH.name}",
        flush=True,
    )


if __name__ == "__main__":
    main()
